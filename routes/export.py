from fastapi import APIRouter
from fastapi.responses import StreamingResponse
from supabase import Client
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import io

router = APIRouter()

def get_router(supabase: Client):

    def generer_excel(lignes, titre):
        wb = Workbook()
        ws = wb.active
        ws.title = titre[:31].replace("/", "-")

        titre_font = Font(bold=True, size=14, color="2d2d2d")
        header_font = Font(bold=True, size=10, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="2d2d2d")
        total_font = Font(bold=True, size=11, color="5d9461")

        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 8
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 12
        ws.column_dimensions['H'].width = 10
        ws.column_dimensions['I'].width = 12
        ws.column_dimensions['J'].width = 12
        ws.column_dimensions['K'].width = 12
        ws.column_dimensions['L'].width = 12

        ws.merge_cells('A1:L1')
        ws['A1'] = titre
        ws['A1'].font = titre_font
        ws['A1'].alignment = Alignment(horizontal="center")

        headers = ['#', 'Article', 'Catégorie', 'Date', 'Qté', 'Achat UV', 'Vente UV', 'Remise %', 'Remise DH', 'Prix net UV', 'CA (DH)', 'Marge (DH)']
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        total_ca = 0
        total_cout = 0
        total_remise = 0
        total_marge = 0

        for i, (l, date_vente) in enumerate(lignes, 1):
            row = i + 3
            produit = l.get("produits", {}) or {}

            nom = l.get("nom_produit", produit.get("nom", ""))
            categorie = l.get("categorie", "")
            qte = l.get("quantite", 0)
            prix_achat = float(l.get("prix_achat", 0))
            prix_vente = float(l.get("prix_vente", 0))
            remise_pct = float(l.get("remise", 0))
            remise_dh = float(l.get("montant_remise", 0))
            ca_ligne = float(l.get("net", 0))                        # CA total de la ligne
            prix_net_unitaire = prix_vente * (1 - remise_pct / 100)  # prix net par unité
            marge = float(l.get("marge", 0))

            ws.cell(row=row, column=1, value=i)
            ws.cell(row=row, column=2, value=nom)
            ws.cell(row=row, column=3, value=categorie)
            ws.cell(row=row, column=4, value=date_vente)
            ws.cell(row=row, column=5, value=qte)
            ws.cell(row=row, column=6, value=prix_achat)
            ws.cell(row=row, column=7, value=prix_vente)
            ws.cell(row=row, column=8, value=remise_pct)
            ws.cell(row=row, column=9, value=remise_dh)
            ws.cell(row=row, column=10, value=prix_net_unitaire)  # Prix net unitaire
            ws.cell(row=row, column=11, value=ca_ligne)           # CA total de la ligne
            ws.cell(row=row, column=12, value=marge)

            total_ca += ca_ligne
            total_cout += prix_achat * qte
            total_remise += remise_dh
            total_marge += marge

        total_row = len(lignes) + 5
        ws.merge_cells(f'A{total_row}:C{total_row}')
        ws.cell(row=total_row, column=1, value="TOTAL").font = total_font

        resume_row = total_row + 2
        resume_data = [
            ("Chiffre d'affaires", total_ca),
            ("Coût marchandises", total_cout),
            ("Total remises accordées", total_remise),
            ("Marge brute", total_marge),
            ("Taux de marge", f"{(total_marge / total_ca * 100):.1f}%" if total_ca > 0 else "0%"),
        ]

        for i, (label, val) in enumerate(resume_data):
            ws.cell(row=resume_row + i, column=1, value=label).font = Font(bold=True)
            cell = ws.cell(row=resume_row + i, column=5, value=val)
            cell.font = total_font
            if isinstance(val, (int, float)):
                cell.number_format = '#,##0.00'

        net_row = resume_row + len(resume_data) + 1
        ws.cell(row=net_row, column=1, value="RÉSULTAT NET").font = Font(bold=True, size=12, color="2d2d2d")
        ws.cell(row=net_row, column=5, value=total_marge).font = Font(bold=True, size=12, color="27ae60")

        # Générer en mémoire — pas d'écriture sur disque
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer

    @router.get("/jour/{date_jour}")
    def export_journee(date_jour: str):
        ventes = supabase.table("ventes").select("*, lignes_vente(*)").eq("date", date_jour).execute()

        lignes = []
        for v in ventes.data:
            for l in v.get("lignes_vente", []):
                lignes.append((l, v.get("date", "")))

        date_formatted = f"{date_jour[8:10]}-{date_jour[5:7]}-{date_jour[0:4]}"
        titre = f"VENTES DU {date_jour[8:10]}/{date_jour[5:7]}/{date_jour[0:4]}"
        buffer = generer_excel(lignes, titre)

        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="Ventes_{date_formatted}.xlsx"'}
        )

    @router.get("/mois/{annee_mois}")
    def export_mois(annee_mois: str):
        ventes = supabase.table("ventes").select("*, lignes_vente(*)") \
            .gte("date", f"{annee_mois}-01") \
            .lte("date", f"{annee_mois}-31") \
            .execute()

        lignes = []
        for v in ventes.data:
            for l in v.get("lignes_vente", []):
                lignes.append((l, v.get("date", "")))

        mois_formatted = f"{annee_mois[5:7]}-{annee_mois[0:4]}"
        titre = f"CUMUL VENTES — MOIS {annee_mois[5:7]}/{annee_mois[0:4]}"
        buffer = generer_excel(lignes, titre)

        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="Ventes_Mois_{mois_formatted}.xlsx"'}
        )

    return router
